#!/usr/bin/env python3
"""
migra_relogios.py
-----------------
Migra os dados da planilha 'BaseCompras' do arquivo Excel de Relógios
para a tabela controle_compras no Supabase.

Uso:
  python migra_relogios.py [--dry-run] [--start-row 3] [--batch-size 50]

Requer as variáveis de ambiente (ou arquivo .env na pasta raiz do projeto):
  VITE_SUPABASE_URL   → URL do projeto Supabase
  VITE_SUPABASE_KEY   → service_role key (não a anon key!)

Obs: "Cancelado" é mapeado para "EXCLUIDO" seguindo o padrão do sistema.
     "Malha ou link?" e "Silicouro?" são combinados em dc_tipo_pulseira.
     Campos REMOVER / None são ignorados.
"""

import sys
import os
import json
import argparse
import datetime
import math
from pathlib import Path

# ---------------------------------------------------------------------------
# Dependências
# ---------------------------------------------------------------------------
try:
    import openpyxl
except ImportError:
    sys.exit("ERRO: instale openpyxl  →  pip install openpyxl")

try:
    import requests
except ImportError:
    sys.exit("ERRO: instale requests  →  pip install requests")

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / '.env.migration')
except ImportError:
    pass  # .env manual

# ---------------------------------------------------------------------------
# Configuração
# ---------------------------------------------------------------------------
SUPABASE_URL = os.environ.get('SUPABASE_URL', "https://vavdakgdtmibajbgcthn.supabase.co")
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', "")

EXCEL_PATH = Path(__file__).parent.parent.parent / 'ControleCompras_RELOGIO - Migração Sysplan.xlsx'
SHEET_NAME = 'BaseCompras'

# Status do Excel → status do sistema
STATUS_MAP = {
    'aberto':    'ABERTO',
    'carteira':  'CARTEIRA',
    'estoque':   'ESTOQUE',
    'cancelado': 'EXCLUIDO',
    'excluido':  'EXCLUIDO',
}

# ---------------------------------------------------------------------------
# Mapeamento de colunas (1-indexed → campo banco)
# ---------------------------------------------------------------------------
def col(n: int):
    """Retorna índice 0-based para acesso à tupla de row."""
    return n - 1

def to_date(v) -> str | None:
    """Converte valor do Excel para string ISO (YYYY-MM-DD) ou None."""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.datetime.strptime(s.split(' ')[0], fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None

def to_str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def to_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def to_int(v) -> int | None:
    f = to_float(v)
    return int(f) if f is not None else None

def anomes(dt_str: str | None) -> int | None:
    """Extrai AnoMes (YYYYMM) de uma data ISO."""
    if not dt_str:
        return None
    return int(dt_str[:4] + dt_str[5:7])

def map_status(v) -> str:
    s = str(v or '').strip().lower()
    return STATUS_MAP.get(s, 'ABERTO')

def combina_pulseira(malha, silicouro) -> str | None:
    """Combina 'Malha ou link?' e 'Silicouro?' em dc_tipo_pulseira."""
    m = to_str(malha)
    s = to_str(silicouro)
    if m and s and s.upper() not in ('NÃO', 'NAO', 'N/A', '-', 'None'):
        return f"{m} / {s}"
    return m or (s if s and s.upper() not in ('NÃO', 'NAO', 'N/A') else None)

def row_to_payload(row: tuple) -> dict:
    """Converte uma linha do Excel num dicionário para inserção no banco."""
    dt_rec = to_date(row[col(47)])
    payload = {
        # Classificação
        'dc_canal':              to_str(row[col(2)]),
        'dc_grupo':              to_str(row[col(3)]),
        'dc_subgrupo':           to_str(row[col(4)]),
        'dc_formato':            to_str(row[col(5)]),
        'dc_sexo':               to_str(row[col(6)]),
        'dc_grupo_planejamento': to_str(row[col(7)]),
        'dc_linha':              to_str(row[col(8)]),
        'dc_griffe':             to_str(row[col(10)]),
        'dc_segmentacao':        to_str(row[col(11)]),
        # Produto
        'dc_material1':          to_str(row[col(13)]),
        'dc_material2':          to_str(row[col(14)]),
        'dc_atributo1':          to_str(row[col(15)]),
        'dc_atributo2':          to_str(row[col(16)]),
        # Campos exclusivos de Relógio
        'dc_tipo_pulseira':      combina_pulseira(row[col(17)], row[col(18)]),
        'dc_tipo_dial':          to_str(row[col(19)]),
        'dc_numeros':            to_str(row[col(20)]),
        'dc_num_maquina':        to_str(row[col(21)]),
        'dc_tamanho':            to_str(row[col(22)]),
        'dc_medidas':            to_str(row[col(23)]),
        'dc_acabamento_caixa':   to_str(row[col(24)]),
        'dc_tipo_visor':         to_str(row[col(25)]),
        'dc_montadora':          to_str(row[col(27)]),
        # Fornecedor / Pedidos
        'dc_fornecedor':         to_str(row[col(26)]),
        'dc_modal':              to_str(row[col(28)]),
        'nr_quantidade':         to_float(row[col(29)]) or 0,
        'nr_fob_negociado':      to_float(row[col(30)]) or 0,
        'nr_total_fob':          to_float(row[col(35)]) or 0,
        'nr_preco_varejo':       to_float(row[col(36)]) or 0,
        # Datas
        'dt_recebimento':        dt_rec,
        'nr_anomes':             anomes(dt_rec),
        # Status / Identificadores
        'dc_status':             map_status(row[col(51)]),
        'cd_codigo_compra':      to_str(row[col(52)]),
        'cd_pedido_fornecedor':  to_str(row[col(53)]),
        'cd_pedido_sap':         to_str(row[col(54)]),
        'cd_spare_parts':        to_str(row[col(55)]),
        'cd_material_pai':       to_str(row[col(56)]),
        'cd_material_fornecedor':to_str(row[col(57)]),
        'dt_delivery':           to_date(row[col(59)]),
        'dt_revised_delivery':   to_date(row[col(60)]),
        'nr_lead_time':          to_int(row[col(61)]),
        'dc_gaveta':             to_str(row[col(64)]),
        'dc_nf_seculus':         to_str(row[col(66)]),
        'dc_observacao':         to_str(row[col(67)]),
        'dc_fup_produto':        to_str(row[col(71)]),
        # Relógio: fob_real = fob_negociado na migração
        'nr_fob_real':           to_float(row[col(30)]) or 0,
    }
    return payload

# ---------------------------------------------------------------------------
# Supabase REST helper
# ---------------------------------------------------------------------------
def supabase_insert(records: list[dict], dry_run: bool) -> tuple[int, int]:
    """Insere registros em lote. Retorna (ok, erros)."""
    if dry_run:
        print(f"  [DRY-RUN] Seria inserido: {len(records)} registro(s)")
        if records:
            print("  Exemplo:", json.dumps(records[0], ensure_ascii=False, default=str)[:300])
        return len(records), 0

    url = f"{SUPABASE_URL}/rest/v1/controle_compras"
    headers = {
        'apikey':        SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type':  'application/json',
        'Prefer':        'return=minimal',
    }
    resp = requests.post(url, headers=headers, json=records, timeout=60)
    if resp.status_code in (200, 201):
        return len(records), 0
    else:
        print(f"  ERRO HTTP {resp.status_code}: {resp.text[:500]}")
        return 0, len(records)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description='Migra relógios para o Supabase')
    parser.add_argument('--dry-run',    action='store_true',  help='Simula sem inserir no banco')
    parser.add_argument('--start-row',  type=int, default=3,  help='Linha inicial de dados (default 3)')
    parser.add_argument('--batch-size', type=int, default=50, help='Tamanho do lote (default 50)')
    args = parser.parse_args()

    if not args.dry_run and (not SUPABASE_URL or not SUPABASE_KEY):
        sys.exit(
            "ERRO: configure VITE_SUPABASE_URL e VITE_SUPABASE_KEY no .env\n"
            "  Use --dry-run para testar sem inserir."
        )

    print(f"📂  Abrindo: {EXCEL_PATH}")
    if not EXCEL_PATH.exists():
        sys.exit(f"ERRO: arquivo não encontrado: {EXCEL_PATH}")

    import shutil
    import tempfile
    temp_file = None
    try:
        wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    except PermissionError:
        print("⚠️  Arquivo bloqueado pelo Excel. Copiando para arquivo temporário...")
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, "relogio_temp_migracao.xlsx")
        shutil.copyfile(str(EXCEL_PATH), temp_file)
        wb = openpyxl.load_workbook(temp_file, read_only=True, data_only=True)

    ws = wb[SHEET_NAME]
    total_rows = ws.max_row - args.start_row + 1
    print(f"📊  Sheet '{SHEET_NAME}' · {total_rows} linhas de dados")
    print(f"{'🔍 DRY-RUN' if args.dry_run else '🚀 INSERÇÃO REAL'} — batch={args.batch_size}")
    print()

    payloads = []
    skipped  = 0
    for i, row in enumerate(ws.iter_rows(min_row=args.start_row, values_only=True), start=args.start_row):
        # Ignora linha completamente vazia
        if not any(row):
            skipped += 1
            continue
        try:
            p = row_to_payload(row)
            payloads.append(p)
        except Exception as e:
            print(f"  ⚠️  Linha {i}: erro no mapeamento — {e}")
            skipped += 1

    wb.close()
    if temp_file and os.path.exists(temp_file):
        try:
            os.remove(temp_file)
        except Exception:
            pass
    print(f"  Linhas válidas: {len(payloads)} · ignoradas: {skipped}")
    print()

    total_ok  = 0
    total_err = 0
    n_batches = math.ceil(len(payloads) / args.batch_size)
    for b in range(n_batches):
        lote = payloads[b * args.batch_size:(b + 1) * args.batch_size]
        print(f"  Lote {b+1}/{n_batches} ({len(lote)} registros)...", end=' ', flush=True)
        ok, err = supabase_insert(lote, args.dry_run)
        total_ok  += ok
        total_err += err
        if not args.dry_run:
            print('✅' if err == 0 else f'❌ {err} erro(s)')

    print()
    print(f"{'='*50}")
    print(f"  Total inseridos : {total_ok}")
    print(f"  Total com erro  : {total_err}")
    if args.dry_run:
        print("  (DRY-RUN — nenhum dado foi inserido no banco)")

if __name__ == '__main__':
    main()
